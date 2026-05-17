"""
Convert the CONUS-AgWeather_v1 station Excel files to Parquet and bundle the
dataset into a single distributable archive.

Steps:
  1. Strip the empty "Filled Data" sheet from every ``*.xlsx`` in
     ``Data/CONUS-AgWeather_v1/standardized_data_xlsx`` (it contains only NaN
     across the entire dataset).
  2. Read every ``*.xlsx`` and write one Parquet file per remaining sheet to
     ``Data/CONUS-AgWeather_v1/standardized_data_parquet``.
     Sheets are mapped to suffixes::
        "Corrected Data"      -> _corrected.parquet
        "Delta (Corr - Orig)" -> _delta.parquet
     Any stale ``*_filled.parquet`` files are removed first.
  3. Zip the entire ``CONUS-AgWeather_v1`` directory (xlsx + parquet + metadata)
     together with ``Plots/Variable_Maps`` into ``CONUS-AgWeather.zip`` at the
     repo root. ``.DS_Store`` files are excluded.

Author: Dr. Sayantan Majumdar (sayantan.majumdar@dri.edu)
"""

import glob
import os
import time
import zipfile
from multiprocessing import Pool

import pandas as pd
from openpyxl import load_workbook


SHEET_TO_SUFFIX = {
    'Corrected Data': 'corrected',
    'Delta (Corr - Orig)': 'delta',
}

DROP_SHEETS = {'Filled Data'}

EXCLUDE_NAMES = {'.DS_Store'}


def _strip_one(in_path):
    """Remove sheets in DROP_SHEETS from an xlsx in place; return (path, removed_any)."""
    try:
        wb = load_workbook(in_path)
        removed = [s for s in DROP_SHEETS if s in wb.sheetnames]
        if not removed:
            wb.close()
            return (in_path, False, None)
        for s in removed:
            del wb[s]
        wb.save(in_path)
        wb.close()
        return (in_path, True, None)
    except Exception as exc:
        return (in_path, False, str(exc))


def strip_filled_sheets(xlsx_dir, n_workers=8):
    xlsx_files = sorted(
        os.path.join(xlsx_dir, f)
        for f in os.listdir(xlsx_dir)
        if f.lower().endswith('.xlsx') and f not in EXCLUDE_NAMES
    )
    if not xlsx_files:
        print(f'No xlsx files found in {xlsx_dir}')
        return

    print(f'Stripping {sorted(DROP_SHEETS)} from {len(xlsx_files)} xlsx files using {n_workers} workers...')
    t0 = time.time()
    with Pool(n_workers) as pool:
        results = pool.map(_strip_one, xlsx_files)
    elapsed = time.time() - t0

    stripped = sum(1 for _, did, err in results if did and err is None)
    skipped = sum(1 for _, did, err in results if not did and err is None)
    errors = [(p, err) for p, _, err in results if err is not None]
    print(f'  done in {elapsed:.1f}s — {stripped} stripped, {skipped} already clean, {len(errors)} errors')
    for p, err in errors[:10]:
        print(f'    error {os.path.basename(p)}: {err}')


def _convert_one(args):
    in_path, out_dir = args
    base = os.path.splitext(os.path.basename(in_path))[0]
    try:
        with pd.ExcelFile(in_path) as xl:
            for sheet in xl.sheet_names:
                suffix = SHEET_TO_SUFFIX.get(
                    sheet, sheet.strip().replace(' ', '_').replace('(', '').replace(')', '').lower()
                )
                df = xl.parse(sheet)
                out_path = os.path.join(out_dir, f'{base}_{suffix}.parquet')
                df.to_parquet(out_path, index=False)
    except Exception as exc:
        print(f'  ! failed {os.path.basename(in_path)}: {exc}')
        return (in_path, False)
    return (in_path, True)


def convert_all(xlsx_dir, parquet_dir, n_workers=8):
    if not os.path.isdir(xlsx_dir):
        raise FileNotFoundError(f'xlsx directory not found: {xlsx_dir}')
    os.makedirs(parquet_dir, exist_ok=True)

    # Remove any stale parquet files for sheets we no longer convert
    stale_suffixes = {'filled'}
    for suffix in stale_suffixes:
        stale = glob.glob(os.path.join(parquet_dir, f'*_{suffix}.parquet'))
        for fp in stale:
            os.remove(fp)
        if stale:
            print(f'  removed {len(stale)} stale *_{suffix}.parquet files')

    xlsx_files = sorted(
        os.path.join(xlsx_dir, f)
        for f in os.listdir(xlsx_dir)
        if f.lower().endswith('.xlsx') and f not in EXCLUDE_NAMES
    )
    if not xlsx_files:
        print(f'No xlsx files found in {xlsx_dir}')
        return

    print(f'Converting {len(xlsx_files)} xlsx files to parquet using {n_workers} workers...')
    t0 = time.time()
    tasks = [(p, parquet_dir) for p in xlsx_files]
    with Pool(n_workers) as pool:
        results = pool.map(_convert_one, tasks)
    elapsed = time.time() - t0

    failed = [p for p, ok in results if not ok]
    succeeded = len(results) - len(failed)
    print(f'  done in {elapsed:.1f}s — {succeeded} succeeded, {len(failed)} failed')
    if failed:
        for p in failed[:10]:
            print(f'    failed: {p}')


def make_zip(src_dirs, zip_path):
    """Zip the given source directories into ``zip_path``.

    Each source directory is added under its basename at the archive root
    (e.g. ``CONUS-AgWeather_v1/...``, ``Variable_Maps/...``). ``.DS_Store``
    files are skipped.
    """
    src_dirs = [os.path.abspath(d) for d in src_dirs]
    print(f'Building {zip_path}')
    for d in src_dirs:
        print(f'  source: {d}')
    print(f'  excluding: {sorted(EXCLUDE_NAMES)}')

    t0 = time.time()
    n_files = 0
    total_bytes = 0
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for src_dir in src_dirs:
            if not os.path.isdir(src_dir):
                print(f'  ! skipping missing source: {src_dir}')
                continue
            parent = os.path.dirname(src_dir)
            for root, dirs, files in os.walk(src_dir):
                dirs[:] = [d for d in dirs if d not in EXCLUDE_NAMES]
                for fname in files:
                    if fname in EXCLUDE_NAMES:
                        continue
                    full = os.path.join(root, fname)
                    arcname = os.path.relpath(full, parent)
                    zf.write(full, arcname)
                    n_files += 1
                    total_bytes += os.path.getsize(full)
    elapsed = time.time() - t0
    out_size = os.path.getsize(zip_path)
    print(f'  wrote {n_files} files ({total_bytes / 1e6:.1f} MB uncompressed -> '
          f'{out_size / 1e6:.1f} MB zip) in {elapsed:.1f}s')


if __name__ == '__main__':
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    data_root = os.path.join(repo_root, 'Data', 'CONUS-AgWeather_v1')
    xlsx_dir = os.path.join(data_root, 'standardized_data_xlsx')
    parquet_dir = os.path.join(data_root, 'standardized_data_parquet')
    zip_path = os.path.join(repo_root, 'CONUS-AgWeather.zip')

    variable_maps_dir = os.path.join(repo_root, 'Plots', 'Variable_Maps')

    strip_filled_sheets(xlsx_dir, n_workers=8)
    convert_all(xlsx_dir, parquet_dir, n_workers=8)
    make_zip([data_root, variable_maps_dir], zip_path)
    print('\nDone.')
